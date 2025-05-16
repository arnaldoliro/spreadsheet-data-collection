import os
from dotenv import load_dotenv

load_dotenv()

def load_data():
    mode = os.getenv("SOURCE_MODE", "google").lower()
    
    if mode == "google":
        return load_google_sheets()
    elif mode == "arquivo":
        return load_local_sheet()
    else:
        raise ValueError("SOURCE_MODE inválido. Use 'google' ou 'arquivo'.")


def load_google_sheets():
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
    client = gspread.authorize(creds)

    sheet_name = os.getenv("GOOGLE_SHEET_NAME")
    tab_name = os.getenv("GOOGLE_WORKSHEET_NAME")

    sheet = client.open(sheet_name).worksheet(tab_name)
    return sheet.get_all_values()


def load_local_sheet():
    from openpyxl import load_workbook

    path = os.getenv("LOCAL_FILE_PATH")
    wb = load_workbook(path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data
